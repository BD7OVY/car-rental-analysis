# -*- coding: utf-8 -*-
"""
租车经营分析平台
============================
把各渠道订单原始 Excel（默认读取 SOURCE_FOLDER，可用 python generate_dashboard.py <文件夹> 覆盖），
归一成统一 schema，计算日/周/月环比 + Top 城市/平台，生成内联 SVG 单文件 HTML。

用法：
    python generate_dashboard.py
输出：
    tools/sample_outputs/分析看板.html
    tools/sample_outputs/metrics.json   （供 server 刷新时读取）
"""
import os
import re
import json
import hashlib
from datetime import datetime, date, timedelta
from collections import defaultdict, Counter

import openpyxl

# ============================================================
# 路径与配置
# ============================================================
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

# 读取配置（数据源 / 输出目录均可被 config.json 覆盖；也可用命令行参数指向任意文件夹）
CONFIG_PATH = os.path.join(HERE, 'config.json')
CONFIG = {}
if os.path.exists(CONFIG_PATH):
    try:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            CONFIG = json.load(f)
    except Exception:
        pass

def _resolve(path, default):
    """config 里的相对路径都相对于项目根 ROOT；绝对路径原样使用；空值回退默认。"""
    p = (path or '').strip()
    if not p:
        return default
    return p if os.path.isabs(p) else os.path.join(ROOT, p)

# 数据源：默认读取 sample-data（各渠道订单 Excel）；命令行参数可指向任意文件夹，不受固定目录约束。
SOURCE_FOLDER = _resolve(CONFIG.get('data_source'), os.path.join(ROOT, 'sample-data'))
# 输出：默认 tools/sample_outputs（分析看板.html / metrics.json / unified_orders.json 等）
OUTPUT_FOLDER = _resolve(CONFIG.get('output_dir'), os.path.join(ROOT, 'tools', 'sample_outputs'))
TODAY = date.today()

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# ============================================================
# 常量：节假日、城市映射、状态映射、车型归一
# ============================================================
HOLIDAYS_2026 = [
    ('元旦',     date(2026, 1, 1),  date(2026, 1, 3)),
    ('春节',     date(2026, 2, 15), date(2026, 2, 23)),
    ('清明节',   date(2026, 4, 4),  date(2026, 4, 6)),
    ('劳动节',   date(2026, 5, 1),  date(2026, 5, 5)),
    ('端午节',   date(2026, 6, 19), date(2026, 6, 21)),
    ('中秋节',   date(2026, 9, 25), date(2026, 9, 27)),
    ('国庆节',   date(2026, 10, 1), date(2026, 10, 7)),
]

# ============================================================
# 加载统一映射配置（单一事实来源：Python 管线与前端导入按钮共用）
# ============================================================
SCHEMA_PATH = os.path.join(ROOT, 'schema_mapping.json')
MAPPING = {}
if os.path.exists(SCHEMA_PATH):
    try:
        with open(SCHEMA_PATH, 'r', encoding='utf-8') as f:
            MAPPING = json.load(f)
    except Exception as e:
        print(f'[警告] 读取 schema_mapping.json 失败: {e}')

STATUS_MAP = MAPPING.get('status_map', {})
MODEL_PATTERNS = MAPPING.get('model_patterns', [])
CITY_MAPS = MAPPING.get('city_maps', {})
HOLIDAYS_2026 = [
    (n, datetime.strptime(s, '%Y-%m-%d').date(), datetime.strptime(e, '%Y-%m-%d').date())
    for n, s, e in MAPPING.get('holidays_2026', [])
]
PLATFORMS = MAPPING.get('platforms', {})

# ============================================================
# 通用解析辅助函数
# ============================================================
def parse_date(val):
    if val is None or val == '' or val == '/':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d', '%Y/%m/%d']:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_datetime(val):
    if val is None or val == '' or val == '/':
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d']:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def fmt_date(d):
    return d.strftime('%Y-%m-%d') if d else None


def parse_gmv(val):
    if val is None or val == '' or val == '/':
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return 0.0


def parse_rental_days(val):
    if val is None or val == '' or val == '/':
        return 0.0
    s = str(val).strip()
    days = 0.0
    m = re.search(r'(\d+)\s*天', s)
    if m:
        days += int(m.group(1))
    m = re.search(r'(\d+)\s*小时', s)
    if m:
        days += int(m.group(1)) / 24.0
    m = re.search(r'(\d+)\s*分钟', s)
    if m:
        days += int(m.group(1)) / 1440.0
    return round(days, 4)


def simplify_model(raw):
    if not raw:
        return ''
    s = str(raw).strip()
    if not s:
        return ''
    for pattern, result in MODEL_PATTERNS:
        if re.search(pattern, s, re.IGNORECASE):
            return result
    # Fallback cleaning
    s = re.sub(r'^\d+-', '', s)
    s = re.sub(r'\s*20\d{2}款.*$', '', s)
    s = re.sub(r'纯电动.*$', '', s)
    s = re.sub(r'磷酸铁锂.*$', '', s)
    s = re.sub(r'三元锂.*$', '', s)
    for prefix in ['广汽埃安', '埃安', '一汽', '东风风行', '东风风神', '东风', '吉利汽车', '吉利', '比亚迪', '北汽新能源', '北汽']:
        if s.startswith(prefix):
            s = s[len(prefix):]
            break
    for kw in ['炫', '魅', '前驱', '网约', '鸿途', '快换', '乐享', 'AIR', '精英', '标准', '豪华', '行政', '经典', '出行', 'PLUS-R', 'Plus-R']:
        idx = s.find(kw)
        if idx > 0:
            s = s[:idx]
    s = re.sub(r'-[A-Z]?\d+.*$', '', s)
    s = s.strip()
    return s if s else '未知车型'


def map_status(raw):
    return STATUS_MAP.get(str(raw).strip() if raw else '', str(raw).strip() if raw else '')


def get_holiday(d):
    if not d:
        return '非节假日'
    for name, start, end in HOLIDAYS_2026:
        if start <= d <= end:
            return name
    return '非节假日'


# ============================================================
# 表头统一：泛型读取器（读取 schema_mapping.json）
#   扫描数据源全部 xlsx → 按列特征签名识别平台 → 映射列名到统一 schema
# ============================================================
def _norm_header(h):
    return re.sub(r'[（）()]', '', str(h if h is not None else '').strip().lower()).replace(' ', '').replace('　', '')


def _build_resolvers():
    """依据 schema 构建：精确候选表 + 分词规则（与前端一致）。"""
    exact = {}
    token_rules = []
    for cfg in PLATFORMS.values():
        for field, cands in (cfg.get('columns') or {}).items():
            if not isinstance(cands, list):
                continue
            for cand in cands:
                nk = _norm_header(cand)
                if nk and nk not in exact:
                    exact[nk] = field
    fk = MAPPING.get('field_keywords', {})
    for field, rules in fk.items():
        for rule in rules:
            score = sum(len(_norm_header(t)) for t in rule)
            token_rules.append({'field': field, 'tokens': [_norm_header(t) for t in rule], 'score': score})
    pri = {'订单号': 0, '实际取车时间': 1, '预计取车时间': 2, '实际还车时间': 3,
           '预计还车时间': 4, '下单时间': 5, '订单金额': 6, '订单状态': 7, '车型': 8,
           '城市列': 9, '车牌号': 10, '租期': 11}
    token_rules.sort(key=lambda t: (-t['score'], pri.get(t['field'], 99)))
    return exact, token_rules


_EXACT, _TOKEN_RULES = _build_resolvers()


def resolve_headers(headers):
    """把一组真实表头解析成 统一字段 → 实际列名（精确候选 → 分词规则）。"""
    hmap = {}
    for h in headers:
        f = _EXACT.get(_norm_header(h))
        if f and f not in hmap:
            hmap[f] = h
    excl = MAPPING.get('field_excludes', {})
    for h in headers:
        nh = _norm_header(h)
        for t in _TOKEN_RULES:
            if all(tok in nh for tok in t['tokens']):
                if excl.get(t['field']) and any(_norm_header(x) in nh for x in excl[t['field']]):
                    continue
                if t['field'] not in hmap:
                    hmap[t['field']] = h
                break
    return hmap


def detect_platform(headers):
    """基于专属标识词(identify)识别平台，不依赖文件名；专属词权重远高于通用列。"""
    nhs = [_norm_header(h) for h in headers]
    best, bs = None, -1
    for plat, cfg in PLATFORMS.items():
        id_hits = 0
        for tok in cfg.get('identify', []):
            tn = _norm_header(tok)
            if any(tn in nh for nh in nhs):
                id_hits += 1
        exact = 0
        for field, cands in (cfg.get('columns') or {}).items():
            if not isinstance(cands, list):
                continue
            for cand in cands:
                if cand in headers:
                    exact += 1
                    break
        sc = id_hits * 1000 + exact
        if sc > bs:
            best, bs = plat, sc
    return best if bs > 0 else None


def read_all_orders():
    """扫描数据源文件夹内全部 xlsx → 探测表头行 → 模糊统一表头 → 按订单号去重。
    支持任意渠道列名变体、标题行偏移、多 sheet；与前端导入引擎口径一致。"""
    orders = []
    seen = set()
    if not os.path.isdir(SOURCE_FOLDER):
        return orders
    for fname in sorted(os.listdir(SOURCE_FOLDER)):
        if not (fname.endswith('.xlsx') and not fname.startswith('~$')):
            continue
        path = os.path.join(SOURCE_FOLDER, fname)
        try:
            wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        except Exception as e:
            print(f'  [跳过] {fname} 读取失败: {e}')
            continue
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            matrix = [list(r) for r in ws.iter_rows(values_only=True)]
            if len(matrix) < 2:
                continue
            # 前 12 行里找「命中分词规则最多」的行作为表头
            hdr_idx, best_score = -1, 0
            for i, row in enumerate(matrix[:12]):
                txt = ' '.join(_norm_header(c) for c in row)
                score = sum(1 for t in _TOKEN_RULES if all(tok in txt for tok in t['tokens']))
                if score > best_score:
                    best_score, hdr_idx = score, i
            if hdr_idx < 0 or best_score < 2:
                continue
            headers = [str(h).strip() if h is not None else '' for h in matrix[hdr_idx]]
            hmap = resolve_headers(headers)
            plat = detect_platform(headers)
            cfg = PLATFORMS.get(plat, {})
            col = cfg.get('columns', {})
            city_map_name = col.get('城市映射')
            city_map = CITY_MAPS.get(city_map_name, {}) if city_map_name else None
            rules = cfg.get('rules', {})
            clear_pickup = set(rules.get('clear_actual_pickup_on', []))
            clear_return = set(rules.get('clear_actual_return_on', []))
            for r in matrix[hdr_idx + 1:]:
                if not any(c is not None and str(c).strip() != '' for c in r):
                    continue
                row = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}

                def g(field):
                    ah = hmap.get(field)
                    if not ah:
                        return None
                    v = row.get(ah)
                    return v if (v is not None and str(v).strip() != '') else None

                oid = str(g('订单号') or '').strip()
                if not oid or oid in seen:
                    continue
                seen.add(oid)
                raw_status = str(g('订单状态') or '').strip()
                status = map_status(raw_status)
                order_time = parse_datetime(g('下单时间'))
                planned_pickup = parse_date(g('预计取车时间'))
                planned_return = parse_date(g('预计还车时间'))
                actual_pickup = parse_date(g('实际取车时间'))
                actual_return = parse_date(g('实际还车时间'))
                if raw_status in clear_pickup:
                    actual_pickup = None
                if raw_status in clear_return:
                    actual_return = None
                gmv = parse_gmv(g('订单金额'))
                rental_days = parse_rental_days(g('租期'))
                raw_model = g('车型')
                model = simplify_model(raw_model)
                raw_city = g('城市列')
                if city_map is not None:
                    city = city_map.get(str(raw_city or '').strip(), '') if raw_city is not None else ''
                else:
                    city = str(raw_city).strip() if raw_city else ''
                plate = str(g('车牌号') or '').strip()
                event_date = actual_pickup or planned_pickup
                holiday = get_holiday(event_date)
                orders.append({
                    '订单号': oid,
                    '下单时间': fmt_date(order_time.date()) if order_time else None,
                    '预计取车时间': fmt_date(planned_pickup),
                    '实际取车时间': fmt_date(actual_pickup),
                    '预计还车时间': fmt_date(planned_return),
                    '实际还车时间': fmt_date(actual_return),
                    '租期': rental_days,
                    '是否节假日订单': '是' if holiday != '非节假日' else '否',
                    '车型': str(raw_model).strip() if raw_model else '',
                    '车型统一': model,
                    '城市': city,
                    '订单金额': gmv,
                    '订单所属平台': plat or '通用',
                    '订单状态': raw_status,
                    '订单状态统一': status,
                    '车牌号': plate,
                })
        wb.close()
    return orders


# ============================================================
# 指标计算
# ============================================================
def period_range(period):
    """返回 (start, end) 日期（含）。period: today/yesterday/this_week/last_week/this_month/last_month"""
    if period == 'today':
        return TODAY, TODAY
    if period == 'yesterday':
        d = TODAY - timedelta(days=1)
        return d, d
    if period == 'this_week':
        start = TODAY - timedelta(days=TODAY.weekday())
        return start, start + timedelta(days=6)
    if period == 'last_week':
        end = TODAY - timedelta(days=TODAY.weekday() + 1)
        start = end - timedelta(days=6)
        return start, end
    if period == 'this_month':
        start = date(TODAY.year, TODAY.month, 1)
        return start, TODAY  # 截至今天
    if period == 'last_month':
        y, m = (TODAY.year, TODAY.month - 1) if TODAY.month > 1 else (TODAY.year - 1, 12)
        start = date(y, m, 1)
        # 上月整月
        end = date(TODAY.year, TODAY.month, 1) - timedelta(days=1)
        return start, end
    return None, None


def filter_by_event_date(orders, start, end):
    """按实际取车时间落在 [start, end] 内过滤；无实际取车时间的不计入。"""
    result = []
    for o in orders:
        d = parse_date(o['实际取车时间'])
        if d and start <= d <= end:
            result.append(o)
    return result


def aggregate(orders):
    """统计订单数、GMV、平均客单价。"""
    gmv = sum(o['订单金额'] for o in orders)
    return {
        'orders': len(orders),
        'gmv': round(gmv, 2),
        'avg_order_value': round(gmv / len(orders), 2) if orders else 0,
    }


def growth(current, previous, key):
    cv = current.get(key, 0)
    pv = previous.get(key, 0)
    if pv == 0:
        return None if cv == 0 else float('inf')
    return round((cv - pv) / pv * 100, 2)


def compute_metrics(orders):
    # 过滤： revenue 相关排除已取消
    valid_orders = [o for o in orders if o['订单状态统一'] != '已取消']

    periods = {}
    for name in ['today', 'yesterday', 'this_week', 'last_week', 'this_month', 'last_month']:
        s, e = period_range(name)
        periods[name] = aggregate(filter_by_event_date(valid_orders, s, e))

    # 环比
    mom_gmv = growth(periods['this_month'], periods['last_month'], 'gmv')
    mom_orders = growth(periods['this_month'], periods['last_month'], 'orders')
    wow_gmv = growth(periods['this_week'], periods['last_week'], 'gmv')
    wow_orders = growth(periods['this_week'], periods['last_week'], 'orders')
    dod_gmv = growth(periods['today'], periods['yesterday'], 'gmv')
    dod_orders = growth(periods['today'], periods['yesterday'], 'orders')

    # 本月数据（按实际取车）
    s, e = period_range('this_month')
    month_orders = filter_by_event_date(valid_orders, s, e)

    # Top 城市 by GMV
    city_gmv = defaultdict(float)
    city_orders = defaultdict(int)
    for o in month_orders:
        if o['城市']:
            city_gmv[o['城市']] += o['订单金额']
            city_orders[o['城市']] += 1
    top_cities = sorted(
        [{'city': c, 'gmv': round(v, 2), 'orders': city_orders[c]} for c, v in city_gmv.items()],
        key=lambda x: x['gmv'], reverse=True)[:10]

    # Top 平台 by GMV
    plat_gmv = defaultdict(float)
    plat_orders = defaultdict(int)
    for o in month_orders:
        p = o['订单所属平台']
        plat_gmv[p] += o['订单金额']
        plat_orders[p] += 1
    top_platforms = sorted(
        [{'platform': p, 'gmv': round(v, 2), 'orders': plat_orders[p]} for p, v in plat_gmv.items()],
        key=lambda x: x['gmv'], reverse=True)

    # 状态分布（本月）
    status_counts = defaultdict(int)
    status_gmv = defaultdict(float)
    for o in month_orders:
        status_counts[o['订单状态统一']] += 1
        status_gmv[o['订单状态统一']] += o['订单金额']
    status_dist = [{'status': s, 'orders': c, 'gmv': round(status_gmv[s], 2)} for s, c in status_counts.items()]

    # 节假日分布（本月）
    holiday_orders = {'是': 0, '否': 0}
    holiday_gmv = {'是': 0.0, '否': 0.0}
    for o in month_orders:
        flag = o['是否节假日订单']
        holiday_orders[flag] = holiday_orders.get(flag, 0) + 1
        holiday_gmv[flag] = holiday_gmv.get(flag, 0.0) + o['订单金额']
    holiday_dist = {
        '是': {'orders': holiday_orders.get('是', 0), 'gmv': round(holiday_gmv.get('是', 0.0), 2)},
        '否': {'orders': holiday_orders.get('否', 0), 'gmv': round(holiday_gmv.get('否', 0.0), 2)},
    }

    # 本月日趋势
    dim = (date(TODAY.year, TODAY.month + 1, 1) - timedelta(days=1)).day if TODAY.month < 12 else 31
    daily = []
    for day in range(1, dim + 1):
        d = date(TODAY.year, TODAY.month, day)
        if d > TODAY:
            break
        day_orders = [o for o in month_orders if parse_date(o['实际取车时间']) == d]
        daily.append({
            'date': d.strftime('%m-%d'),
            'full_date': d.strftime('%Y-%m-%d'),
            'orders': len(day_orders),
            'gmv': round(sum(o['订单金额'] for o in day_orders), 2),
        })

    # 关键车型（按本月 GMV）
    model_gmv = defaultdict(float)
    for o in month_orders:
        if o['车型统一']:
            model_gmv[o['车型统一']] += o['订单金额']
    top_models = sorted([{'model': m, 'gmv': round(g, 2)} for m, g in model_gmv.items()],
                        key=lambda x: x['gmv'], reverse=True)[:5]

    metrics = {
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'today_str': TODAY.strftime('%Y-%m-%d'),
        'periods': periods,
        'growth': {
            'month_over_month_gmv': mom_gmv,
            'month_over_month_orders': mom_orders,
            'week_over_week_gmv': wow_gmv,
            'week_over_week_orders': wow_orders,
            'day_over_day_gmv': dod_gmv,
            'day_over_day_orders': dod_orders,
        },
        'top_cities': top_cities,
        'top_platforms': top_platforms,
        'top_models': top_models,
        'status_distribution': status_dist,
        'holiday_distribution': holiday_dist,
        'daily_trend': daily,
        'total_orders_all': len(orders),
        'total_orders_valid': len(valid_orders),
    }

    # 结构化洞察（harness：按清单硬编码模板，非 LLM 自由发挥）
    insights = []
    p = periods
    g = metrics['growth']
    def fmt_pct(v):
        if v is None: return '—'
        if v == float('inf'): return '新增'
        return f"{v:+.1f}%"
    insights.append(
        f"本月（截至 {TODAY.strftime('%m月%d日')}）实际取车 {p['this_month']['orders']} 单，"
        f"GMV ¥{p['this_month']['gmv']:,.0f}，环比上月 {fmt_pct(g['month_over_month_gmv'])}。"
    )
    insights.append(
        f"本周（{period_range('this_week')[0].strftime('%m.%d')}-{period_range('this_week')[1].strftime('%m.%d')}）"
        f"GMV ¥{p['this_week']['gmv']:,.0f}，环比上周 {fmt_pct(g['week_over_week_gmv'])}。"
    )
    insights.append(
        f"今日 GMV ¥{p['today']['gmv']:,.0f}，较昨日 {fmt_pct(g['day_over_day_gmv'])}。"
    )
    if top_cities:
        top3 = '、'.join([f"{x['city']}({x['gmv']:,.0f})" for x in top_cities[:3]])
        insights.append(f"本月 Top3 城市：{top3}。")
    if top_platforms:
        top3p = '、'.join([f"{x['platform']}({x['gmv']:,.0f})" for x in top_platforms[:3]])
        insights.append(f"本月 Top3 平台：{top3p}。")
    h = holiday_dist
    if h['是']['orders']:
        insights.append(
            f"本月节假日订单 {h['是']['orders']} 单 / ¥{h['是']['gmv']:,.0f}，"
            f"占比 {h['是']['orders']/(h['是']['orders']+h['否']['orders'])*100:.1f}。"
        )
    # 集中度 / 异常观察（始终输出）
    month_total = p['this_month']['gmv']
    if month_total and top_cities:
        share = top_cities[0]['gmv'] / month_total * 100
        insights.append(
            f"收入集中度：{top_cities[0]['city']} 单城贡献本月 {share:.0f}% GMV，需关注单一城市依赖风险。"
        )
    peak = max(daily, key=lambda d: d['gmv']) if daily else None
    if peak and peak['gmv'] > 0:
        insights.append(f"本月峰值日 {peak['date']} 单日 GMV ¥{peak['gmv']:,.0f}，为日常均值的 {peak['gmv']/max(month_total/(p['this_month']['orders'] or 1),1):.0f} 倍。")
    if h['是']['orders'] == 0 and (h['否']['orders'] + h['是']['orders']) > 0:
        insights.append("本月无节假日订单，节假日溢价红利尚未释放。")
    metrics['insights'] = insights
    return metrics


# ============================================================
# 内联 SVG 图表生成
# ============================================================
_SVG_UID = 0

def svg_line_chart(daily, width=640, height=240):
    """本月 GMV 日趋势折线图（带渐变面积填充）。"""
    global _SVG_UID
    if not daily:
        return ''
    _SVG_UID += 1
    gid = f"lg{_SVG_UID}"
    vals = [d['gmv'] for d in daily]
    max_v = max(vals) if vals else 1
    max_v = max(max_v, 1)
    pad_l, pad_r, pad_t, pad_b = 44, 20, 28, 40
    w = width - pad_l - pad_r
    h = height - pad_t - pad_b
    n = len(daily)
    step = w / (n - 1) if n > 1 else w

    pts = []
    for i, d in enumerate(daily):
        x = pad_l + i * step
        y = pad_t + h - (d['gmv'] / max_v * h)
        pts.append((x, y))
    poly_points = ' '.join(f"{x:.1f},{y:.1f}" for x, y in pts)
    last_x = pad_l + (n - 1) * step
    area = (f"M {pad_l:.1f} {pad_t + h:.1f} "
            + " L ".join(f"{x:.1f} {y:.1f}" for x, y in pts)
            + f" L {last_x:.1f} {pad_t + h:.1f} Z")

    # Y 轴刻度
    ticks = []
    for i in range(5):
        v = max_v * i / 4
        y = pad_t + h - (i / 4 * h)
        ticks.append(f'<text x="{pad_l-8}" y="{y+4:.1f}" text-anchor="end" font-size="10" fill="#64748b">¥{v/1000:.0f}k</text>')
        ticks.append(f'<line x1="{pad_l}" y1="{y:.1f}" x2="{width-pad_r}" y2="{y:.1f}" stroke="#eef2f7" stroke-width="1"/>')

    # X 轴标签（隔一天显示）
    xlabels = []
    for i, d in enumerate(daily):
        if i % 2 == 0 or i == n - 1:
            x = pad_l + i * step
            xlabels.append(f'<text x="{x:.1f}" y="{height-12}" text-anchor="middle" font-size="10" fill="#64748b">{d["date"]}</text>')

    dots = ''.join(
        f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="#2563eb" stroke="#fff" stroke-width="1.5"/>'
        for x, y in pts
    )

    svg = f'''<svg viewBox="0 0 {width} {height}" style="width:100%;height:auto;">
  <defs><linearGradient id="{gid}" x1="0" y1="0" x2="0" y2="1">
    <stop offset="0%" stop-color="#2563eb" stop-opacity="0.28"/>
    <stop offset="100%" stop-color="#2563eb" stop-opacity="0.02"/>
  </linearGradient></defs>
  {' '.join(ticks)}
  <path d="{area}" fill="url(#{gid})" stroke="none"/>
  <polyline points="{poly_points}" fill="none" stroke="#2563eb" stroke-width="2.5" stroke-linejoin="round" stroke-linecap="round"/>
  {dots}
  {' '.join(xlabels)}
</svg>'''
    return svg


CAT_COLORS = ['#2563eb', '#16a34a', '#f59e0b', '#8b5cf6', '#06b6d4', '#ef4444',
               '#0ea5e9', '#84cc16', '#f43f5e', '#a855f7']

def svg_horizontal_bar(data, label_key, value_key, width=500, height=260, colors=None):
    """横向条形图。data 为 list[dict]，colors 为分类配色列表。"""
    if not data:
        return ''
    palette = colors or CAT_COLORS
    max_v = max(d[value_key] for d in data) or 1
    pad_l, pad_r, pad_t, pad_b = 90, 30, 20, 20
    w = width - pad_l - pad_r
    h = height - pad_t - pad_b
    n = len(data)
    bar_h = h / n * 0.6
    gap = h / n * 0.4

    rects = []
    labels = []
    vals = []
    for i, d in enumerate(data):
        y = pad_t + i * (bar_h + gap) + gap / 2
        bw = max(d[value_key] / max_v * w, 2)  # 至少 2px 可见
        color = palette[i % len(palette)]
        rects.append(f'<rect x="{pad_l}" y="{y:.1f}" width="{bw:.1f}" height="{bar_h:.1f}" rx="5" fill="{color}"/>')
        labels.append(f'<text x="{pad_l-8}" y="{y+bar_h/2+4:.1f}" text-anchor="end" font-size="12" fill="#334155">{d[label_key]}</text>')
        # 数值放在条内（若条够长）或条外
        if bw > 54:
            vals.append(f'<text x="{pad_l+bw-8:.1f}" y="{y+bar_h/2+4:.1f}" text-anchor="end" font-size="11" font-weight="600" fill="#fff">¥{d[value_key]:,.0f}</text>')
        else:
            vals.append(f'<text x="{pad_l+bw+6:.1f}" y="{y+bar_h/2+4:.1f}" font-size="11" fill="#475569">¥{d[value_key]:,.0f}</text>')

    svg = f'''<svg viewBox="0 0 {width} {height}" style="width:100%;height:auto;">
  {' '.join(rects)}
  {' '.join(labels)}
  {' '.join(vals)}
</svg>'''
    return svg


def svg_donut(data, label_key, value_key, width=280, height=240):
    """环形图。data 为 list[dict]，value 是订单数。"""
    if not data:
        return ''
    total = sum(d[value_key] for d in data)
    if total == 0:
        return ''
    colors = ['#2563eb', '#16a34a', '#f59e0b', '#ef4444', '#8b5cf6', '#06b6d4']
    cx, cy = width / 2, height / 2
    r_out, r_in = 80, 50
    start = -90  # 从顶部开始
    arcs = []
    legend = []
    for i, d in enumerate(data):
        frac = d[value_key] / total
        angle = frac * 360
        end = start + angle
        # SVG arc flag
        large = 1 if angle > 180 else 0
        x1 = cx + r_out * 3.1415926 / 180 * (start)  # approximate? better use math.cos/sin
        # use real trig below
        import math
        rad1 = math.radians(start)
        rad2 = math.radians(end)
        x1 = cx + r_out * math.cos(rad1)
        y1 = cy + r_out * math.sin(rad1)
        x2 = cx + r_out * math.cos(rad2)
        y2 = cy + r_out * math.sin(rad2)
        x3 = cx + r_in * math.cos(rad2)
        y3 = cy + r_in * math.sin(rad2)
        x4 = cx + r_in * math.cos(rad1)
        y4 = cy + r_in * math.sin(rad1)
        color = colors[i % len(colors)]
        path = f'M {x1:.1f} {y1:.1f} A {r_out} {r_out} 0 {large} 1 {x2:.1f} {y2:.1f} L {x3:.1f} {y3:.1f} A {r_in} {r_in} 0 {large} 0 {x4:.1f} {y4:.1f} Z'
        arcs.append(f'<path d="{path}" fill="{color}"/>')
        pct = frac * 100
        legend.append(
            f'<rect x="{width-110}" y="{30 + i*20}" width="10" height="10" rx="2" fill="{color}"/>'
            f'<text x="{width-95}" y="{40 + i*20}" font-size="11" fill="#334155">{d[label_key]} {pct:.0f}%</text>'
        )
        start = end

    svg = f'''<svg viewBox="0 0 {width} {height}" style="width:100%;height:auto;">
  {' '.join(arcs)}
  <text x="{cx}" y="{cy+5}" text-anchor="middle" font-size="16" font-weight="600" fill="#1e293b">{total}</text>
  <text x="{cx}" y="{cy+22}" text-anchor="middle" font-size="10" fill="#64748b">单</text>
  {' '.join(legend)}
</svg>'''
    return svg


def svg_period_compare(metrics, width=640, height=260):
    """环比对比分组柱状图：本月/本周/今日 × 本期 vs 上期 GMV。"""
    p = metrics['periods']
    g = metrics['growth']
    groups = [
        ('本月', 'this_month', 'last_month', g.get('month_over_month_gmv')),
        ('本周', 'this_week', 'last_week', g.get('week_over_week_gmv')),
        ('今日', 'today', 'yesterday', g.get('day_over_day_gmv')),
    ]
    pad_l, pad_r, pad_t, pad_b = 44, 20, 36, 34
    w = width - pad_l - pad_r
    h = height - pad_t - pad_b
    gw = w / len(groups)              # 每组宽度
    inner_pad = gw * 0.22
    bar_w = (gw - inner_pad * 2) / 2  # 每组两根柱
    max_v = max(max(p[c]['gmv'], p[pr]['gmv']) for _, c, pr, _ in groups) or 1

    parts = []
    # 基准横线
    parts.append(f'<line x1="{pad_l}" y1="{pad_t+h:.1f}" x2="{width-pad_r}" y2="{pad_t+h:.1f}" stroke="#e2e8f0" stroke-width="1"/>')
    for gi, (name, cur, prev, pct) in enumerate(groups):
        cx0 = pad_l + gi * gw
        cv, pv = p[cur]['gmv'], p[prev]['gmv']
        ch = cv / max_v * h
        ph = pv / max_v * h
        x_cur = cx0 + inner_pad
        x_prev = cx0 + inner_pad + bar_w
        parts.append(f'<rect x="{x_cur:.1f}" y="{pad_t+h-ch:.1f}" width="{bar_w:.1f}" height="{ch:.1f}" rx="3" fill="#2563eb"/>')
        parts.append(f'<rect x="{x_prev:.1f}" y="{pad_t+h-ph:.1f}" width="{bar_w:.1f}" height="{ph:.1f}" rx="3" fill="#cbd5e1"/>')
        # 增长百分比
        if pct is None:
            ptxt = '—'
            pcol = '#64748b'
        elif pct == float('inf'):
            ptxt = '新增'
            pcol = '#16a34a'
        else:
            ptxt = f"{pct:+.0f}%"
            pcol = '#16a34a' if pct >= 0 else '#dc2626'
        midx = cx0 + gw / 2
        parts.append(f'<text x="{midx:.1f}" y="{pad_t-16:.1f}" text-anchor="middle" font-size="13" font-weight="700" fill="{pcol}">{ptxt}</text>')
        parts.append(f'<text x="{midx:.1f}" y="{pad_t-3:.1f}" text-anchor="middle" font-size="11" fill="#475569">{name}</text>')
        # 数值标签
        if ch > 14:
            parts.append(f'<text x="{x_cur+bar_w/2:.1f}" y="{pad_t+h-ch-4:.1f}" text-anchor="middle" font-size="9" fill="#1e40af">¥{cv/1000:.1f}k</text>')
        if ph > 14:
            parts.append(f'<text x="{x_prev+bar_w/2:.1f}" y="{pad_t+h-ph-4:.1f}" text-anchor="middle" font-size="9" fill="#64748b">¥{pv/1000:.1f}k</text>')
    # 图例
    parts.append(f'<rect x="{pad_l}" y="{height-14}" width="10" height="10" rx="2" fill="#2563eb"/>'
                 f'<text x="{pad_l+15}" y="{height-5}" font-size="10" fill="#475569">本期</text>')
    parts.append(f'<rect x="{pad_l+70}" y="{height-14}" width="10" height="10" rx="2" fill="#cbd5e1"/>'
                 f'<text x="{pad_l+85}" y="{height-5}" font-size="10" fill="#475569">上期</text>')
    svg = f'''<svg viewBox="0 0 {width} {height}" style="width:100%;height:auto;">
  {''.join(parts)}
</svg>'''
    return svg


# ============================================================
# HTML 看板生成
# ============================================================
def render_html(metrics):
    p = metrics['periods']
    g = metrics['growth']

    def fmt_money(v):
        return f"¥{v:,.0f}" if isinstance(v, (int, float)) else '¥0'

    def fmt_num(v):
        return f"{int(v)}" if isinstance(v, (int, float)) else '0'

    def arrow(v):
        if v is None: return '<span class="flat">—</span>'
        if v == float('inf'): return '<span class="up">新增</span>'
        cls = 'up' if v > 0 else 'down'
        return f'<span class="{cls}">{v:+.1f}%</span>'

    def trend_color(v):
        if v is None: return '#64748b'
        return '#16a34a' if v > 0 else '#dc2626'

    line_chart = svg_line_chart(metrics['daily_trend'], width=640, height=240)
    city_chart = svg_horizontal_bar(metrics['top_cities'], 'city', 'gmv', width=500, height=280)
    plat_chart = svg_horizontal_bar(metrics['top_platforms'], 'platform', 'gmv', width=500, height=200)
    status_chart = svg_donut(metrics['status_distribution'], 'status', 'orders', width=280, height=220)
    compare_chart = svg_period_compare(metrics, width=640, height=260)

    insights_html = ''.join([f'<li>{txt}</li>' for txt in metrics['insights']])

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>租车经营分析平台</title>
<style>
:root {{
  --bg:#f8fafc; --card:#fff; --text:#1e293b; --muted:#64748b; --border:#e2e8f0;
  --blue:#2563eb; --green:#16a34a; --red:#dc2626; --amber:#f59e0b;
}}
*{{box-sizing:border-box;}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,"Helvetica Neue",Arial,sans-serif;background:var(--bg);color:var(--text);margin:0;padding:16px;line-height:1.5;}}
.container{{max-width:1200px;margin:0 auto;}}
header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;flex-wrap:wrap;gap:10px;}}
h1{{font-size:22px;margin:0;}}
.meta{{color:var(--muted);font-size:13px;}}
.btn{{display:inline-flex;align-items:center;gap:6px;padding:10px 16px;border-radius:8px;border:1px solid var(--border);background:#fff;color:var(--text);font-size:14px;cursor:pointer;min-height:44px;}}
.btn:hover{{background:#f1f5f9;}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:16px;margin-bottom:20px;}}
.card{{background:var(--card);border-radius:12px;padding:16px;box-shadow:0 1px 2px rgba(0,0,0,0.04);border:1px solid var(--border);}}
.card h3{{margin:0 0 6px;font-size:13px;color:var(--muted);font-weight:500;}}
.card .big{{font-size:26px;font-weight:700;margin-bottom:4px;}}
.card .sub{{font-size:12px;color:var(--muted);}}
.up{{color:var(--green);font-weight:600;}}
.down{{color:var(--red);font-weight:600;}}
.flat{{color:var(--muted);}}
.section-title{{font-size:16px;font-weight:600;margin:24px 0 12px;}}
.charts{{display:grid;grid-template-columns:repeat(auto-fit,minmax(320px,1fr));gap:16px;}}
.chart-card{{background:var(--card);border-radius:12px;padding:16px;border:1px solid var(--border);}}
.chart-card h3{{margin:0 0 12px;font-size:14px;color:var(--muted);font-weight:500;}}
.insight{{background:#eff6ff;border:1px solid #bfdbfe;border-radius:12px;padding:14px 18px;margin-bottom:20px;}}
.insight-title{{font-size:15px;font-weight:700;color:#1d4ed8;margin-bottom:8px;}}
.insight ul{{margin:0;padding-left:18px;color:#1e3a8a;}}
.insight li{{margin:4px 0;}}
table{{width:100%;border-collapse:collapse;font-size:13px;}}
th,td{{text-align:left;padding:8px;border-bottom:1px solid var(--border);}}
th{{color:var(--muted);font-weight:500;}}
tr:last-child td{{border-bottom:none;}}
.footer{{margin-top:24px;color:var(--muted);font-size:12px;text-align:center;}}
@media (max-width:768px){{
  body{{padding:12px;}}
  h1{{font-size:18px;}}
  .grid{{grid-template-columns:1fr;}}
  .charts{{grid-template-columns:1fr;}}
}}
</style>
</head>
<body>
<div class="container">
  <header>
    <div>
      <h1>租车经营分析平台</h1>
      <div class="meta">数据截至 {metrics['today_str']} · 按实际取车时间统计 · 已排除已取消订单</div>
    </div>
    <div style="display:flex;gap:10px;flex-wrap:wrap;align-items:center;">
      <button class="btn" id="refreshBtn" onclick="refreshData()">刷新数据</button>
      <button class="btn" onclick="window.print()">打印 / 另存 PDF</button>
      <span id="refreshTip" style="font-size:12px;color:#64748b;"></span>
    </div>
  </header>

  <div class="insight">
    <div class="insight-title">数据洞察</div>
    <ul>{insights_html}</ul>
  </div>

  <div class="section-title">核心指标</div>
  <div class="grid">
    <div class="card">
      <h3>本月 GMV</h3>
      <div class="big">{fmt_money(p['this_month']['gmv'])}</div>
      <div class="sub">环比 {arrow(g['month_over_month_gmv'])} · {fmt_num(p['this_month']['orders'])} 单</div>
    </div>
    <div class="card">
      <h3>本周 GMV</h3>
      <div class="big">{fmt_money(p['this_week']['gmv'])}</div>
      <div class="sub">环比 {arrow(g['week_over_week_gmv'])} · {fmt_num(p['this_week']['orders'])} 单</div>
    </div>
    <div class="card">
      <h3>今日 GMV</h3>
      <div class="big">{fmt_money(p['today']['gmv'])}</div>
      <div class="sub">环比 {arrow(g['day_over_day_gmv'])} · {fmt_num(p['today']['orders'])} 单</div>
    </div>
    <div class="card">
      <h3>累计有效订单</h3>
      <div class="big">{fmt_num(metrics['total_orders_valid'])}</div>
      <div class="sub">原始数据共 {fmt_num(metrics['total_orders_all'])} 单</div>
    </div>
  </div>

  <div class="section-title">环比对比</div>
  <div class="charts">
    <div class="chart-card" style="grid-column:span 2;">
      <h3>GMV 环比：本期 vs 上期（本月 / 本周 / 今日）</h3>
      {compare_chart}
    </div>
  </div>

  <div class="section-title">趋势与分布</div>
  <div class="charts">
    <div class="chart-card" style="grid-column:span 2;">
      <h3>本月 GMV 日趋势</h3>
      {line_chart}
    </div>
    <div class="chart-card">
      <h3>平台 GMV 对比</h3>
      {plat_chart}
    </div>
    <div class="chart-card">
      <h3>本月订单状态分布</h3>
      {status_chart}
    </div>
    <div class="chart-card" style="grid-column:span 2;">
      <h3>城市 GMV Top10</h3>
      {city_chart}
    </div>
  </div>

  <div class="section-title">城市 Top10 明细</div>
  <div class="card">
    <table>
      <thead><tr><th>排名</th><th>城市</th><th>订单数</th><th>GMV</th></tr></thead>
      <tbody>
        {''.join([f'<tr><td>{i+1}</td><td>{x["city"]}</td><td>{x["orders"]}</td><td>{fmt_money(x["gmv"])}</td></tr>' for i,x in enumerate(metrics['top_cities'])])}
      </tbody>
    </table>
  </div>

  <div class="section-title">平台明细</div>
  <div class="card">
    <table>
      <thead><tr><th>平台</th><th>订单数</th><th>GMV</th><th>客单价</th></tr></thead>
      <tbody>
        {''.join([f'<tr><td>{x["platform"]}</td><td>{x["orders"]}</td><td>{fmt_money(x["gmv"])}</td><td>{fmt_money(x["gmv"]/x["orders"] if x["orders"] else 0)}</td></tr>' for x in metrics['top_platforms']])}
      </tbody>
    </table>
  </div>

  <div class="footer">生成时间：{metrics['generated_at']} · 租车经营分析工作台</div>
</div>
<script>
function refreshData() {{
  var btn = document.getElementById('refreshBtn');
  var tip = document.getElementById('refreshTip');
  if (tip) {{ tip.textContent = '刷新中…'; tip.style.color = '#f59e0b'; }}
  fetch('/api/refresh', {{method:'POST'}})
    .then(function(r){{ return r.json(); }})
    .then(function(d){{
      if (d.ok) {{
        if (tip) {{ tip.textContent = '已刷新：' + d.message; tip.style.color = '#16a34a'; }}
        setTimeout(function(){{ location.reload(); }}, 700);
      }} else {{
        if (tip) {{ tip.textContent = '刷新失败：' + d.message; tip.style.color = '#dc2626'; }}
      }}
    }})
    .catch(function(e){{
      if (tip) {{ tip.textContent = '刷新需通过本地服务打开（直接双击 HTML 无效）'; tip.style.color = '#dc2626'; }}
    }});
}}
</script>
</body>
</html>'''
    return html


# ============================================================
# 主流程
# ============================================================
def main():
    print('=== 租车经营分析平台 · 看板生成 ===')
    print(f'数据源: {SOURCE_FOLDER}')
    print(f'输出目录: {OUTPUT_FOLDER}')
    print(f'今天: {TODAY}')
    print()

    if not os.path.isdir(SOURCE_FOLDER):
        print(f'[错误] 数据源不存在: {SOURCE_FOLDER}')
        return

    print('读取并统一各平台数据（按列特征自动识别平台）...')
    all_orders = read_all_orders()
    pc = Counter(o['订单所属平台'] for o in all_orders)
    print('  ' + '，'.join(f'{k} {v} 单' for k, v in pc.items()) if pc else '  (未识别到任何平台订单)')
    print(f'  合计 {len(all_orders)} 单')
    print()

    print('计算指标...')
    metrics = compute_metrics(all_orders)

    # 写出指标 JSON（server 刷新用）
    metrics_path = os.path.join(OUTPUT_FOLDER, 'metrics.json')
    with open(metrics_path, 'w', encoding='utf-8') as f:
        json.dump(metrics, f, ensure_ascii=False, indent=2)
    print(f'指标已保存: {metrics_path}')

    # 写出统一订单明细（可选，供后续复用）
    unified_path = os.path.join(OUTPUT_FOLDER, 'unified_orders.json')
    with open(unified_path, 'w', encoding='utf-8') as f:
        json.dump(all_orders, f, ensure_ascii=False, indent=2)
    print(f'统一订单明细已保存: {unified_path}')

    # 同步导出前端 data.json（根目录，供 index.html 加载）
    data_json_path = os.path.join(ROOT, 'data.json')
    with open(data_json_path, 'w', encoding='utf-8') as f:
        json.dump(all_orders, f, ensure_ascii=False, indent=2)
    print(f'前端数据已保存: {data_json_path}')

    # 生成 HTML
    html = render_html(metrics)
    html_path = os.path.join(OUTPUT_FOLDER, '分析看板.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'看板已生成: {html_path}')

    print()
    print(f'Total: {len(all_orders)} orders')


if __name__ == '__main__':
    import sys
    # 允许命令行指定任意数据源文件夹（不被 config.json 的 data_source 约束）：
    #   python generate_dashboard.py "D:/某渠道订单"
    if len(sys.argv) > 1 and os.path.isdir(sys.argv[1]):
        SOURCE_FOLDER = sys.argv[1]
        print(f'数据源已被命令行参数覆盖为: {SOURCE_FOLDER}')
    main()
